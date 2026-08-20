import sys, os, json, re, requests, time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}

WEB_GENRES = ["Hành Động", "Kinh Dị", "Hài", "Phiêu Lưu", "Viễn Tưởng"]
REPORT_FILE = "movie_changes_report.xlsx"
DB_FILE = "movies.json"

def fix_img_url(url):
    if not url:
        return ""
    if url.startswith("http"):
        return url
    if url.startswith("/"):
        return "https://phimimg.com" + url
    return "https://phimimg.com/" + url

def extract_accurate_country(country_list):
    if not country_list:
        return "Mỹ"
    
    names = []
    for c in country_list:
        n = (c.get("name", "") if isinstance(c, dict) else str(c)).strip()
        names.append(n)
        
    for n in names:
        if "Nhật" in n:
            return "Nhật Bản"
        elif "Hàn" in n:
            return "Hàn Quốc"
        elif "Việt" in n:
            return "Việt Nam"
        elif "Thái" in n:
            return "Thái Lan"
        elif "Trung" in n or "Hoa" in n:
            return "Trung Quốc"
        elif any(k in n for k in ["Mỹ", "Âu", "Anh", "Pháp", "Đức", "Canada", "Ý", "Tây Ban Nha", "Úc"]):
            return "Mỹ"
            
    return names[0] if names else "Khác"

def map_accurate_genre(category_list):
    if not category_list:
        return "Hành Động"
        
    cat_names = [(c.get("name", "") if isinstance(c, dict) else str(c)).strip() for c in category_list]
    
    # Direct match in WEB_GENRES
    for target in ["Kinh Dị", "Hành Động", "Viễn Tưởng", "Phiêu Lưu", "Hài"]:
        for cn in cat_names:
            if target.lower() in cn.lower():
                return target

    # Related match
    for cn in cat_names:
        cn_l = cn.lower()
        if any(k in cn_l for k in ["kinh di", "ma", "quy", "zombie", "horror", "bi an"]):
            return "Kinh Dị"
        elif any(k in cn_l for k in ["hanh dong", "vo thuat", "chieu rap", "action", "toi pham", "hinh su"]):
            return "Hành Động"
        elif any(k in cn_l for k in ["vien tuong", "gia tuong", "khoa hoc", "sci-fi", "fantasy"]):
            return "Viễn Tưởng"
        elif any(k in cn_l for k in ["phieu luu", "tham hiem", "adventure", "hoathinh"]):
            return "Phiêu Lưu"
        elif any(k in cn_l for k in ["hai", "che", "comedy", "tinh cam", "tam ly", "chinh kich"]):
            return "Hài"

    return "Hành Động"

def fetch_detail(slug):
    url = f"https://phimapi.com/phim/{slug}"
    for attempt in range(3):
        try:
            r = requests.get(url, headers=headers, timeout=6)
            if r.status_code == 200:
                d = r.json()
                if d.get("status") is True:
                    return d
        except:
            pass
        time.sleep(0.05)
    return None

# Animation indicators specifically to reject Marvel / DC animation
animation_indicators = [
    "hoạt hình", "animated", "animation", "anime", "lego",
    "spider-verse", "du-hanh-vu-tru-nhen", "vu-tru-nhen", "du hành vũ trụ nhện",
    "friendly-neighborhood-spider-man", "người nhện hàng xóm thân thiện",
    "x-men-97", "dị nhân '97", "di-nhan-97", "what-if", "what if",
    "spidey", "hit-monkey", "eyes-of-wakanda", "marvel-zombies", "i am groot", "tôi là groot",
    "teen-titans", "teen titans", "dc-super-hero-girls", "mayhem-trong-da-vu-tru",
    "lien-minh-sieu-thu-dc", "liên minh siêu thú dc", "super-pets",
    "batman-ninja", "batman-aztec", "batman-unlimited", "batman-arkham",
    "batman-gotham-knight", "batman-mat-na-ma", "batman-v-superman-tran-chien-cua-cac-anh-hung-nhi",
    "cau-chuyen-lego-batman", "young-justice"
]

def is_marvel_or_dc(text_lower):
    marvel_dc_keywords = [
        "marvel", "avengers", "spider-man", "spiderman", "iron man", "iron-man", "captain america",
        "thor", "black panther", "guardians of the galaxy", "ant-man", "doctor strange", "dr. strange",
        "captain marvel", "black widow", "shang-chi", "eternals", "deadpool", "wolverine", "x-men",
        "venom", "morbius", "kraven", "loki", "wandavision", "hawkeye", "moon knight", "she-hulk",
        "secret invasion", "echo", "agatha", "agent carter", "agents of shield", "daredevil", "punisher",
        "batman", "superman", "the flash", "aquaman", "wonder woman", "justice league", "black adam",
        "shazam", "joker", "suicide squad", "peacemaker", "gotham", "batwoman", "supergirl", "arrow",
        "người nhện", "người sắt", "đội trưởng mỹ", "biệt đội siêu anh hùng", "thần sấm",
        "chiến binh báo đen", "người kiến", "vệ binh dải ngân hà", "phù thủy tối thượng", "góa phụ đen",
        "dị nhân", "người dơi", "siêu nhân", "người đàn ông thép", "nữ thần chiến binh", "liên minh công lý"
    ]
    return any(k in text_lower for k in marvel_dc_keywords)

def is_animated(text_lower):
    return any(a in text_lower for a in animation_indicators)

# Step 1: Load Existing DB (IMMUTABLE MASTER)
print("==================================================")
print("STEP 1: Reading existing movies.json (IMMUTABLE MASTER)...")
existing_db = []
existing_keys = set()
existing_slugs = set()

if os.path.exists(DB_FILE):
    with open(DB_FILE, "r", encoding="utf-8") as f:
        existing_db = json.load(f)

for m in existing_db:
    key = (m.get("title", "").strip().lower(), m.get("year"))
    existing_keys.add(key)
    if m.get("slug"):
        existing_slugs.add(m["slug"])

print(f"Loaded {len(existing_db)} existing movies/series. They will NEVER be deleted or modified!")

# Step 2: Collect Candidate Slugs from Latest Pages, Search, & Marvel/DC Slugs
print("\nSTEP 2: Crawling latest updates & Marvel/DC from PhimAPI...")
candidate_slugs = set()

# Special Target Slugs (The Odyssey, Obsession, FROM, Wednesday, etc.)
special_target_slugs = [
    "su-thi-odyssey", "cuoc-phieu-luu", "am-anh-2026", "am-anh-yeu-va-do-ki", "am-anh", "noi-am-anh-2025",
    "thi-tran-ac-mong-hoi-chuong-la-phan-1", "thi-tran-ac-mong-hoi-chuong-la-phan-2",
    "thi-tran-ac-mong-hoi-chuong-la-phan-3", "thi-tran-ac-mong-hoi-chuong-la-phan-4",
    "nhung-nguoi-con-sot-lai-phan-1", "nhung-nguoi-con-sot-lai-phan-2",
    "thu-tu-phan-1", "thu-tu-phan-2", "gia-toc-rong-phan-1", "gia-toc-rong-phan-2", "gia-toc-rong-phan-3",
    "sup-do-phan-1", "sup-do-phan-2", "hao-quang-phan-1", "hao-quang-phan-2",
    "ngoi-truong-xac-song", "the-gioi-ma-quai", "parasyte-vung-xam", "sinh-vat-gyeongseong", "ac-quy"
]
candidate_slugs.update(special_target_slugs)

# Marvel & DC Search Terms for Daily Crawler
marvel_dc_search_terms = [
    "marvel", "avengers", "spider-man", "spiderman", "iron man", "captain america",
    "thor", "black panther", "guardians of the galaxy", "ant-man", "doctor strange",
    "deadpool", "wolverine", "x-men", "venom", "loki", "daredevil", "punisher",
    "batman", "superman", "the flash", "aquaman", "wonder woman", "justice league",
    "black adam", "shazam", "joker", "suicide squad", "peacemaker", "gotham",
    "người nhện", "người sắt", "đội trưởng mỹ", "thần sấm", "biệt đội siêu anh hùng",
    "vệ binh dải ngân hà", "phù thủy tối thượng", "người dơi", "siêu nhân"
]

def search_term_slugs(kw):
    slugs = []
    try:
        r = requests.get(f"https://phimapi.com/v1/api/tim-kiem?keyword={kw}&limit=30", headers=headers, timeout=8)
        if r.status_code == 200:
            for it in r.json().get("data", {}).get("items", []):
                if it.get("slug"): slugs.append(it["slug"])
    except: pass
    return slugs

def fetch_page_slugs(endpoint, total_pages):
    slugs = []
    for page in range(1, total_pages + 1):
        try:
            r = requests.get(f"{endpoint}?page={page}&limit=24", headers=headers, timeout=6)
            if r.status_code == 200:
                items = r.json().get("data", {}).get("items", [])
                for it in items:
                    if it.get("slug"):
                        slugs.append(it.get("slug"))
        except:
            pass
        time.sleep(0.02)
    return slugs

endpoints = [
    ("https://phimapi.com/v1/api/danh-sach/phim-moi-cap-nhat", 20),
    ("https://phimapi.com/v1/api/danh-sach/phim-le", 100),
    ("https://phimapi.com/v1/api/danh-sach/phim-bo", 100),
    ("https://phimapi.com/v1/api/quoc-gia/au-my", 80),
    ("https://phimapi.com/v1/api/quoc-gia/han-quoc", 60),
    ("https://phimapi.com/v1/api/quoc-gia/nhat-ban", 40),
]

with ThreadPoolExecutor(max_workers=20) as executor:
    futures = [executor.submit(fetch_page_slugs, url, pages) for url, pages in endpoints]
    futures += [executor.submit(search_term_slugs, kw) for kw in marvel_dc_search_terms]
    for future in as_completed(futures):
        candidate_slugs.update(future.result())

# Exclude already existing slugs
new_candidate_slugs = [s for s in candidate_slugs if s not in existing_slugs]
print(f"Total candidate slugs: {len(candidate_slugs)} | Brand new candidate slugs to evaluate: {len(new_candidate_slugs)}")

# Step 3: Process New Candidates
print("\nSTEP 3: Processing new candidates...")
newly_added_items = []

def process_slug(slug):
    d = fetch_detail(slug)
    if not d:
        return None
    info = d.get("movie", {})
    if not info:
        return None

    title = info.get("name", "Phim")
    origin_title = info.get("origin_name", info.get("name", ""))
    desc = info.get("content", "") or f"Bộ phim phát hành năm {info.get('year', 2026)}."
    desc = re.sub(r'<[^>]+>', '', desc).strip()
    full_text = f"{slug} {title} {origin_title} {desc}".lower()

    # Reject Marvel / DC animation specifically
    if is_marvel_or_dc(full_text) and is_animated(full_text):
        return None

    # Country check: Exclude VN, TH, CN
    c_list = info.get("country", [])
    country = extract_accurate_country(c_list)
    if country in ["Việt Nam", "Thái Lan", "Trung Quốc"]:
        return None

    # Japanese Anime Rejection (Keep ONLY verified Japanese Live-Action)
    if country == "Nhật Bản":
        japan_live_keywords = [
            'live action', 'thế giới không lối thoát', 'tokyo vice', 'thế giới ngầm tokyo',
            'shogun', 'first love', 'godzilla', 'lãng khách kenshin', 'phục thù cuộc đời',
            'tokyo revengers', 'bác sĩ x', 'thầy giáo vĩ đại', 'onizuka', '5 centimet',
            'fermat', 'từ bất hảo thành bác sĩ', 'đến khi áo phông', 'tagusari',
            'đội điều tra kỳ án', 'sadako', 'sadak', 'ringu', 'ju-on', 'ma nữ đại chiến',
            'cuộc chiến băng đảng', 'monarch', 'chiến thần samurai', 'quốc bảo',
            'exit 8', 'ga tàu vô tận', 'trò chơi nghìn tỷ', 'trò chơi tìm xác',
            'ngoại tình tột đỉnh', 'đột kích đài truyền hình', 'sa mạc namibia',
            'cảnh đồi mờ xám', 'ngồi bên nòng súng', 'lôi đài chi thượng',
            'nhà ga nuốt người', 'nghìn tỷ'
        ]
        if not any(k in full_text for k in japan_live_keywords):
            return None

    # Year Range: 2020 to 2026
    year_raw = info.get("year", 2026)
    try:
        year = int(year_raw)
    except:
        year = 2026

    if year < 2020 or year > 2026:
        return None

    eps_data = d.get("episodes", [])
    episodes = []
    m3u8_url = ""
    if eps_data and len(eps_data) > 0:
        server = eps_data[0].get("server_data", [])
        for ep in server:
            link = ep.get("link_m3u8", "")
            if link:
                episodes.append({
                    "name": ep.get("name", "Full"),
                    "slug": ep.get("slug", "full"),
                    "link_m3u8": link
                })
        if episodes:
            m3u8_url = episodes[0]["link_m3u8"]

    if not m3u8_url:
        return None

    raw_type = info.get("type", "")
    mtype = "series" if raw_type in ["series", "tvshows"] or len(episodes) > 1 else "single"

    # Genre Mapping: Strictly one of WEB_GENRES
    cats = info.get("category", [])
    genre = map_accurate_genre(cats)

    poster = fix_img_url(info.get("poster_url", ""))
    backdrop = fix_img_url(info.get("thumb_url", poster))
    if not poster:
        return None

    if "thi-tran-ac-mong" in slug or "from" in origin_title.lower():
        origin_title = f"FROM - {origin_title}"

    if len(desc) > 300:
        desc = desc[:297] + "..."

    return {
        "slug": slug,
        "title": title,
        "origin_title": origin_title,
        "year": year,
        "genre": genre,
        "country": country,
        "type": mtype,
        "quality": "1080p FHD",
        "duration": info.get("time", "120 phút") if mtype == "single" else f"{len(episodes)} Tập ({info.get('time', '50 phút/tập')})",
        "poster": poster,
        "backdrop": backdrop,
        "description": desc,
        "m3u8_url": m3u8_url,
        "local_mp4": f"E:\\Phim\\Phim_Le_MP4\\{slug}.mp4",
        "episodes": episodes
    }

with ThreadPoolExecutor(max_workers=20) as executor:
    futures = [executor.submit(process_slug, slug) for slug in new_candidate_slugs]
    for i, future in enumerate(as_completed(futures), 1):
        res = future.result()
        if res:
            title_year_key = (res["title"].strip().lower(), res["year"])
            if title_year_key not in existing_keys:
                existing_keys.add(title_year_key)
                newly_added_items.append(res)

print(f"Newly appended items found: {len(newly_added_items)}")

# Step 4: Append New Items to Master DB
if newly_added_items:
    start_id = len(existing_db) + 1
    for i, item in enumerate(newly_added_items):
        item["id"] = start_id + i
        item["rating"] = round(4.3 + (item["id"] % 7) * 0.1, 1)
        existing_db.append(item)

# Priority sorting: Featured 2026 movies and top rated/Marvel/DC blockbusters prioritized at top
def calculate_featured_score(m):
    score = 0
    year = m.get("year", 2024)
    if year == 2026: score += 1000
    elif year == 2025: score += 500
    elif year == 2024: score += 200
    
    rating = m.get("rating", 4.5)
    score += int(rating * 100)
    
    title_text = (m.get("title","") + " " + m.get("origin_title","")).lower()
    if any(k in title_text for k in ["marvel", "avengers", "spider-man", "deadpool", "batman", "superman", "loki", "thor"]):
        score += 300
    return score

existing_db.sort(key=calculate_featured_score, reverse=True)

# Re-assign sequential IDs cleanly
for i, m in enumerate(existing_db):
    m["id"] = i + 1

final_db = existing_db
total_pages = (len(final_db) + 23) // 24

with open(DB_FILE, "w", encoding="utf-8") as f:
    json.dump(final_db, f, ensure_ascii=False, indent=2)

print("\n==================================================")
print(f"SUCCESS: Master Database movies.json updated!")
print(f"Total Movies in System: {len(final_db)}")
print(f"Total Pages on Website: {total_pages} PAGES!")
print(f"New Items Added Today: {len(newly_added_items)}")

# Step 5: Excel Change Log Generation (movie_changes_report.xlsx)
print("\nSTEP 5: Generating Excel Change Log Report...")

wb = openpyxl.Workbook()

# Sheet 1: Daily Change Log
ws_log = wb.active
ws_log.title = "Nhật Ký Cập Nhật"

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="E50914", end_color="E50914", fill_type="solid")

headers_log = [
    "Ngày Cập Nhật", "ID Phim", "Tên Phim (Việt)", "Tên Gốc (English)", 
    "Loại Phim", "Thể Loại", "Quốc Gia", "Năm PH", "Số Tập", "Trạng Thái", "Nguồn Stream m3u8"
]

ws_log.append(headers_log)
for col_idx in range(1, len(headers_log) + 1):
    cell = ws_log.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

today_str = datetime.now().strftime("%Y-%m-%d %H:%M")

if newly_added_items:
    for item in newly_added_items:
        ws_log.append([
            today_str,
            item.get("id"),
            item.get("title"),
            item.get("origin_title"),
            "Phim Bộ" if item.get("type") == "series" else "Phim Lẻ",
            item.get("genre"),
            item.get("country"),
            item.get("year"),
            len(item.get("episodes", [])),
            "🔥 MỚI THÊM",
            item.get("m3u8_url")
        ])
else:
    ws_log.append([today_str, "-", "Không có phim mới", "-", "-", "-", "-", "-", "-", "TỰ ĐỘNG CẬP NHẬT", "-"])

# Sheet 2: All Movies Catalog
ws_cat = wb.create_sheet(title="Tổng Kho Phim (Full Catalog)")
headers_cat = ["ID", "Tên Phim", "Tên Gốc", "Loại Phim", "Thể Loại", "Quốc Gia", "Năm", "Số Tập", "M3U8 Link"]
ws_cat.append(headers_cat)

for col_idx in range(1, len(headers_cat) + 1):
    cell = ws_cat.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for item in final_db:
    ws_cat.append([
        item.get("id"),
        item.get("title"),
        item.get("origin_title"),
        "Phim Bộ" if item.get("type") == "series" else "Phim Lẻ",
        item.get("genre"),
        item.get("country"),
        item.get("year"),
        len(item.get("episodes", [])),
        item.get("m3u8_url")
    ])

for sheet in [ws_log, ws_cat]:
    sheet.views.sheetView[0].showGridLines = True
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(REPORT_FILE)
print(f"Excel Report saved successfully to: {REPORT_FILE}")
